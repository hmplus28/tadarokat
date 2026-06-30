import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook

from config import DATA_DIR, LOCAL_EXCEL_PATH, LOCAL_SHEETS, PURCHASE_EDITABLE, STORAGE_BACKEND

LOCAL_EXCEL_PATH = LOCAL_EXCEL_PATH  # re-export

PURCHASE_EDIT_HEADERS = ["شماره", *PURCHASE_EDITABLE, "overrides_json", "updated_at", "updated_by"]
INQUIRY_HEADERS = [
    "شماره استعلام", "شماره درخواست خرید", "نوع خرید", "تاریخ استعلام", "وضعیت",
    "مهلت استعلام", "واحد/رمز تامین", "رمز فوریت", "علت خرید", "انبار",
    "درخواست دهنده", "ریسک عدم خرید", "شماره درخواست کالا", "تاریخ درخواست کالا",
    "تاریخ دریافت", "صادر کننده سند", "کارشناس خرید",
    "created_at", "created_by",
]
PREINVOICE_HEADERS = [
    "id", "شماره استعلام", "شماره پیش فاکتور", "نام پیمانکار", "شهر پیمانکار", "تاریخ پیش فاکتور",
    "نوع فاکتور", "شرح",
    "مالیات بر ارزش افزوده", "اعمال مالیات ده درصد", "تخفیف", "زمان تحویل", "توضیحات", "جمع کل",
    "انتخاب شده", "وضعیت مدیر", "کامنت مدیر", "تاریخ بررسی", "بررسی کننده", "created_at", "created_by",
]
LINE_HEADERS = [
    "id", "preinvoice_id", "ردیف", "عنوان کالا", "واحد", "فی", "تعداد", "جمع کل", "توضیحات",
    "منتخب مدیر", "کارشناس ارجاع", "شماره دستور",
]
PRODUCT_HISTORY_HEADERS = [
    "id", "کد قلم خریدنی", "عنوان کالا", "فی", "تعداد", "واحد",
    "نام پیمانکار", "شهر پیمانکار", "شماره استعلام", "شماره خرید",
    "تاریخ خرید", "created_at", "created_by",
]

ORDER_HEADERS = [
    "id", "شماره دستور", "شماره استعلام", "شماره خرید", "ردیف", "عنوان کالا", "واحد", "فی", "تعداد",
    "انبار", "کارشناس", "نام پیمانکار", "وضعیت", "مرحله فعلی", "تاریخ دستور",
    "شماره سفارش", "تاریخ سفارش",
    "شماره پرداخت", "تاریخ ثبت پرداخت",
    "تاریخ واریز",
    "شماره مجوز ورود", "تاریخ تحویل",
    "شماره تحویل", "شماره رسید", "تاریخ رسید",
    "توضیحات", "صادر کننده",
    "created_at", "created_by", "updated_at",
]
DELIVERY_HEADERS = [
    "id", "شماره تحویل", "شماره دستور", "شماره خرید", "عنوان کالا", "انبار",
    "مقدار", "واحد", "تاریخ تحویل", "تحویل گیرنده", "وضعیت", "توضیحات",
    "created_at", "created_by", "updated_at",
]
NOTIFICATION_HEADERS = [
    "id", "username", "warehouse", "عنوان", "پیام", "نوع", "مرجع", "خوانده شده", "created_at",
]
EDIT_HISTORY_HEADERS = [
    "id", "نوع موجودیت", "شناسه", "عملیات", "فیلد", "مقدار قبلی", "مقدار جدید", "کاربر", "created_at",
]

SHEET_HEADERS = {
    "purchase_edits": PURCHASE_EDIT_HEADERS,
    "issued_inquiries": INQUIRY_HEADERS,
    "pre_invoices": PREINVOICE_HEADERS,
    "pre_invoice_lines": LINE_HEADERS,
    "product_history": PRODUCT_HISTORY_HEADERS,
    "orders": ORDER_HEADERS,
    "deliveries": DELIVERY_HEADERS,
    "notifications": NOTIFICATION_HEADERS,
    "edit_history": EDIT_HISTORY_HEADERS,
}


def _normalize_id(val: Any) -> str:
    s = str(val).strip()
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        return s[:-2]
    return s


def _ensure_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _new_id() -> str:
    return str(uuid.uuid4())[:12]


def _use_sqlite() -> bool:
    return STORAGE_BACKEND == "sqlite"


def ensure_local_workbook() -> None:
    _ensure_dir()
    if _use_sqlite():
        from services.db_storage import ensure_database
        ensure_database()
        return
    if not LOCAL_EXCEL_PATH.exists():
        wb = Workbook()
        first = True
        for key, headers in SHEET_HEADERS.items():
            if first:
                ws = wb.active
                ws.title = LOCAL_SHEETS[key]
                first = False
            else:
                ws = wb.create_sheet(LOCAL_SHEETS[key])
            ws.append(headers)
        wb.save(LOCAL_EXCEL_PATH)
        return

    wb = load_workbook(LOCAL_EXCEL_PATH)
    changed = False
    for key, headers in SHEET_HEADERS.items():
        title = LOCAL_SHEETS[key]
        if title not in wb.sheetnames:
            ws = wb.create_sheet(title)
            ws.append(headers)
            changed = True
    if changed:
        wb.save(LOCAL_EXCEL_PATH)


def _normalize_col_name(name: str) -> str:
    return " ".join(str(name).replace("\n", " ").replace("\r", " ").split()).strip()


def _align_dataframe(df: pd.DataFrame, headers: List[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=headers)
    df = df.copy()
    df.columns = [_normalize_col_name(c) for c in df.columns]
    for col in headers:
        if col not in df.columns:
            df[col] = None
    return df


def _read_local_sheet(sheet_key: str) -> pd.DataFrame:
    if _use_sqlite():
        from services.db_storage import read_sheet
        return read_sheet(sheet_key)
    ensure_local_workbook()
    sheet_name = LOCAL_SHEETS[sheet_key]
    headers = SHEET_HEADERS[sheet_key]
    try:
        df = pd.read_excel(LOCAL_EXCEL_PATH, sheet_name=sheet_name, engine="openpyxl")
        return _align_dataframe(df, headers)
    except Exception:
        return pd.DataFrame(columns=headers)


def _write_sheet(sheet_key: str, df: pd.DataFrame) -> None:
    if _use_sqlite():
        from services.db_storage import write_sheet
        write_sheet(sheet_key, df)
        return
    ensure_local_workbook()
    sheet_name = LOCAL_SHEETS[sheet_key]
    headers = SHEET_HEADERS[sheet_key]
    out = _align_dataframe(df, headers)
    with pd.ExcelWriter(
        LOCAL_EXCEL_PATH,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        out[headers].to_excel(writer, sheet_name=sheet_name, index=False)


UPSERT_KEY_COLUMNS = {
    "issued_inquiries": "شماره استعلام",
    "pre_invoices": "id",
    "pre_invoice_lines": "id",
    "orders": "id",
    "deliveries": "id",
    "purchase_edits": "شماره",
    "product_history": "id",
}


def upsert_sheet_rows(sheet_key: str, records: List[Dict[str, Any]], username: str) -> Dict[str, int]:
    """درج یا به‌روزرسانی ردیف‌های import — بر اساس کلید یکتا هر شیت."""
    if sheet_key not in SHEET_HEADERS:
        raise ValueError(f"شیت نامعتبر: {sheet_key}")
    headers = SHEET_HEADERS[sheet_key]
    key_col = UPSERT_KEY_COLUMNS.get(sheet_key, "id")
    now = datetime.utcnow().isoformat()
    df = _read_local_sheet(sheet_key)
    if df.empty:
        df = pd.DataFrame(columns=headers)
    inserted = 0
    updated = 0
    skipped = 0

    for raw in records:
        if not raw or not any(v is not None and str(v).strip() not in ("", "nan", "None") for v in raw.values()):
            skipped += 1
            continue
        row = {h: None for h in headers}
        for k, v in raw.items():
            nk = _normalize_col_name(str(k))
            for h in headers:
                if _normalize_col_name(h) == nk or h == k:
                    row[h] = v
                    break
        if key_col == "id":
            if not row.get("id"):
                row["id"] = _new_id()
        if not row.get("created_at"):
            row["created_at"] = now
        if not row.get("created_by"):
            row["created_by"] = username
        if sheet_key in ("orders", "deliveries") and not row.get("updated_at"):
            row["updated_at"] = now

        key_val = row.get(key_col)
        if key_val is None or str(key_val).strip() == "":
            if key_col == "id":
                row["id"] = row.get("id") or _new_id()
                key_val = row["id"]
            else:
                skipped += 1
                continue

        key_str = str(key_val).strip()
        if key_col in df.columns and not df.empty:
            mask = df[key_col].astype(str).str.strip() == key_str
            if mask.any():
                idx = df[mask].index[0]
                for k, v in row.items():
                    if v is not None and str(v).strip() not in ("", "nan"):
                        if df[k].dtype != object:
                            df[k] = df[k].astype(object)
                        df.at[idx, k] = v
                if sheet_key in ("orders", "deliveries"):
                    df.at[idx, "updated_at"] = now
                updated += 1
                continue

        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
        inserted += 1

    _write_sheet(sheet_key, df)
    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def get_purchase_edits() -> pd.DataFrame:
    df = _read_local_sheet("purchase_edits")
    if df.empty or "شماره" not in df.columns:
        return pd.DataFrame(columns=PURCHASE_EDIT_HEADERS)
    df["شماره"] = df["شماره"].astype(str)
    return df.drop_duplicates(subset=["شماره"], keep="last")


def get_issued_inquiries() -> pd.DataFrame:
    return _read_local_sheet("issued_inquiries")


def get_pre_invoices() -> pd.DataFrame:
    return _read_local_sheet("pre_invoices")


def get_pre_invoice_lines() -> pd.DataFrame:
    return _read_local_sheet("pre_invoice_lines")


def _split_purchase_edit_fields(fields: Dict[str, Any], admin: bool = False) -> tuple:
    import json
    from config import PURCHASE_EDITABLE, PURCHASE_EDIT_BLOCKED

    standard: Dict[str, Any] = {}
    overrides: Dict[str, Any] = {}
    for key, val in fields.items():
        if key in PURCHASE_EDIT_BLOCKED:
            continue
        if val is None:
            continue
        sval = str(val).strip()
        if sval == "" or sval.lower() in ("nan", "none"):
            continue
        if key in PURCHASE_EDITABLE:
            standard[key] = val
        elif admin:
            overrides[key] = val
    return standard, overrides


def save_purchase_edit(
    request_number: str,
    fields: Dict[str, Any],
    username: str,
    admin: bool = False,
) -> Dict[str, Any]:
    import json

    standard, overrides = _split_purchase_edit_fields(fields, admin=admin)
    if _use_sqlite():
        now = datetime.utcnow().isoformat()
        df = get_purchase_edits()
        mask = df["شماره"].astype(str) == str(request_number) if not df.empty else pd.Series([], dtype=bool)
        if not df.empty and mask.any():
            idx = df[mask].index[0]
            for k, v in standard.items():
                df.at[idx, k] = v
            if overrides:
                prev: Dict[str, Any] = {}
                raw_prev = df.at[idx, "overrides_json"] if "overrides_json" in df.columns else None
                if raw_prev:
                    try:
                        prev = json.loads(raw_prev) if isinstance(raw_prev, str) else dict(raw_prev)
                    except Exception:
                        prev = {}
                prev.update(overrides)
                df.at[idx, "overrides_json"] = json.dumps(prev, ensure_ascii=False)
            df.at[idx, "updated_at"] = now
            df.at[idx, "updated_by"] = username
        else:
            row = {h: None for h in PURCHASE_EDIT_HEADERS}
            row["شماره"] = str(request_number)
            row["updated_at"] = now
            row["updated_by"] = username
            for k, v in standard.items():
                row[k] = v
            if overrides:
                row["overrides_json"] = json.dumps(overrides, ensure_ascii=False)
            df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
        _write_sheet("purchase_edits", df)
        match = df[df["شماره"].astype(str) == str(request_number)]
        return match.iloc[0].to_dict() if not match.empty else row

    ensure_local_workbook()
    wb = load_workbook(LOCAL_EXCEL_PATH)
    ws = wb[LOCAL_SHEETS["purchase_edits"]]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    now = datetime.utcnow().isoformat()
    allowed = standard

    target_row = None
    num_col = headers.index("شماره") + 1
    for row_idx in range(2, ws.max_row + 1):
        if str(ws.cell(row=row_idx, column=num_col).value) == str(request_number):
            target_row = row_idx
            break

    if target_row is None:
        row_data = {h: None for h in headers}
        row_data["شماره"] = str(request_number)
        row_data["updated_at"] = now
        row_data["updated_by"] = username
        for k, v in allowed.items():
            if k in headers:
                row_data[k] = v
        ws.append([row_data.get(h) for h in headers])
    else:
        for field, value in allowed.items():
            if field in headers:
                ws.cell(row=target_row, column=headers.index(field) + 1, value=value)
        ws.cell(row=target_row, column=headers.index("updated_at") + 1, value=now)
        ws.cell(row=target_row, column=headers.index("updated_by") + 1, value=username)

    wb.save(LOCAL_EXCEL_PATH)
    return {"request_number": request_number, "updated": allowed}


def append_inquiry(record: Dict[str, Any], username: str) -> Dict[str, Any]:
    now = datetime.utcnow().isoformat()
    row = {
        "شماره استعلام": record.get("شماره استعلام"),
        "شماره درخواست خرید": record.get("شماره درخواست خرید"),
        "نوع خرید": record.get("نوع خرید"),
        "تاریخ استعلام": record.get("تاریخ استعلام"),
        "وضعیت": record.get("وضعیت", "ثبت شده"),
        "مهلت استعلام": record.get("مهلت استعلام"),
        "واحد/رمز تامین": record.get("واحد/رمز تامین", "تدارکات داخلی"),
        "رمز فوریت": record.get("رمز فوریت"),
        "علت خرید": record.get("علت خرید"),
        "انبار": record.get("انبار"),
        "درخواست دهنده": record.get("درخواست دهنده"),
        "ریسک عدم خرید": record.get("ریسک عدم خرید"),
        "شماره درخواست کالا": record.get("شماره درخواست کالا"),
        "تاریخ درخواست کالا": record.get("تاریخ درخواست کالا"),
        "تاریخ دریافت": record.get("تاریخ دریافت"),
        "صادر کننده سند": record.get("صادر کننده سند"),
        "کارشناس خرید": record.get("کارشناس خرید"),
        "created_at": now,
        "created_by": username,
    }
    df = get_issued_inquiries()
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    _write_sheet("issued_inquiries", df)
    return row


def append_pre_invoice(record: Dict[str, Any], username: str) -> Dict[str, Any]:
    now = datetime.utcnow().isoformat()
    row = {
        "id": record.get("id") or _new_id(),
        "شماره استعلام": record.get("شماره استعلام"),
        "شماره پیش فاکتور": record.get("شماره پیش فاکتور"),
        "نام پیمانکار": record.get("نام پیمانکار"),
        "شهر پیمانکار": record.get("شهر پیمانکار"),
        "تاریخ پیش فاکتور": record.get("تاریخ پیش فاکتور"),
        "نوع فاکتور": record.get("نوع فاکتور") or ("رسمی" if record.get("اعمال مالیات ده درصد") else "کد ملی"),
        "شرح": record.get("شرح"),
        "مالیات بر ارزش افزوده": record.get("مالیات بر ارزش افزوده", 0),
        "اعمال مالیات ده درصد": bool(record.get("اعمال مالیات ده درصد")),
        "تخفیف": record.get("تخفیف", 0),
        "زمان تحویل": record.get("زمان تحویل"),
        "توضیحات": record.get("توضیحات"),
        "جمع کل": record.get("جمع کل", 0),
        "انتخاب شده": record.get("انتخاب شده", False),
        "وضعیت مدیر": record.get("وضعیت مدیر", "در انتظار"),
        "کامنت مدیر": record.get("کامنت مدیر"),
        "تاریخ بررسی": None,
        "بررسی کننده": None,
        "created_at": now,
        "created_by": username,
    }
    df = get_pre_invoices()
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    _write_sheet("pre_invoices", df)
    return row


def append_pre_invoice_lines(lines: List[Dict[str, Any]], preinvoice_id: str) -> List[Dict]:
    df = get_pre_invoice_lines()
    rows = []
    for line in lines:
        qty = float(line.get("تعداد") or 0)
        price = float(line.get("فی") or 0)
        total = float(line.get("جمع کل") or qty * price)
        row_num = line.get("ردیف") or line.get("row") or line.get("شماره خرید")
        rows.append({
            "id": line.get("id") or _new_id(),
            "preinvoice_id": preinvoice_id,
            "ردیف": row_num,
            "عنوان کالا": line.get("عنوان کالا"),
            "واحد": line.get("واحد"),
            "فی": price,
            "تعداد": qty,
            "جمع کل": total,
            "توضیحات": line.get("توضیحات"),
            "منتخب مدیر": None,
            "کارشناس ارجاع": None,
            "شماره دستور": None,
        })
    if rows:
        df = pd.concat([df, pd.DataFrame(rows)], ignore_index=True)
        _write_sheet("pre_invoice_lines", df)
    return rows


def _ensure_preinvoice_columns(df: pd.DataFrame) -> pd.DataFrame:
    for col in PREINVOICE_HEADERS:
        if col not in df.columns:
            df[col] = None
    return df


def _expert_in_inquiry_row(row: pd.Series, expert: str) -> bool:
    if not expert:
        return True
    expert = str(expert).strip()
    for col in ("کارشناس خرید", "صادر کننده سند"):
        val = str(row.get(col) or "")
        if expert in val:
            return True
    return False


def _inquiries_for_purchase(purchase_number: str, expert: str = "") -> pd.DataFrame:
    inq_df = get_issued_inquiries()
    if inq_df.empty or "شماره درخواست خرید" not in inq_df.columns:
        return pd.DataFrame(columns=INQUIRY_HEADERS)
    key = _normalize_id(purchase_number)
    match = inq_df[inq_df["شماره درخواست خرید"].map(_normalize_id) == key]
    if expert:
        match = match[match.apply(lambda r: _expert_in_inquiry_row(r, expert), axis=1)]
    return match


def purchase_has_approved_inquiry(purchase_number: str, expert: str = "") -> bool:
    match = _inquiries_for_purchase(purchase_number, expert)
    if match.empty:
        return False
    for _, inq in match.iterrows():
        inq_num = str(inq.get("شماره استعلام", ""))
        if inquiry_has_orders(inq_num):
            return True
    return False


def purchase_has_local_inquiry(purchase_number: str, expert: str = "") -> bool:
    return not _inquiries_for_purchase(purchase_number, expert).empty


def get_local_inquiry_by_purchase(purchase_number: str, expert: str = "") -> Optional[str]:
    match = _inquiries_for_purchase(purchase_number, expert)
    if match.empty:
        return None
    return str(match.iloc[-1].get("شماره استعلام", ""))


def _normalize_order_number(val: Any) -> str:
    s = str(val or "").strip()
    if not s or s.lower() in ("none", "nan"):
        return ""
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        return s[:-2]
    return s


def line_has_order(line_id: str) -> bool:
    lines_df = get_pre_invoice_lines()
    if lines_df.empty:
        return False
    match = lines_df[lines_df["id"].astype(str) == str(line_id)]
    if match.empty:
        return False
    line = match.iloc[0]
    order_num = _normalize_order_number(line.get("شماره دستور"))
    if not order_num:
        return False
    orders_df = get_orders()
    if orders_df.empty:
        return False
    order_nums = orders_df["شماره دستور"].astype(str).map(_normalize_order_number)
    return order_num in order_nums.values


def _normalize_row_key(val: Any) -> str:
    if val is None:
        return ""
    text = str(val).strip()
    if not text or text.lower() in ("nan", "none"):
        return ""
    try:
        return str(int(float(text)))
    except (TypeError, ValueError):
        return text


def get_inquiry_line_order_stats(inquiry_number: str) -> Dict[str, Any]:
    pre_df = get_pre_invoices()
    lines_df = get_pre_invoice_lines()
    empty = {
        "total_lines": 0,
        "total_rows": 0,
        "selected_lines": 0,
        "lines_with_orders": 0,
        "pending_order_lines": 0,
        "pending_row_decisions": 0,
        "pending_line_ids": [],
    }
    if pre_df.empty or lines_df.empty:
        return empty

    pre_ids = pre_df[pre_df["شماره استعلام"].astype(str) == str(inquiry_number)]["id"].astype(str).tolist()
    if not pre_ids:
        return empty

    inq_lines = lines_df[lines_df["preinvoice_id"].astype(str).isin(pre_ids)].copy()
    if inq_lines.empty:
        return empty

    def _truthy(val: Any) -> bool:
        if val is True:
            return True
        return str(val or "").strip().lower() in ("true", "1", "1.0", "yes")

    inq_lines["_row_key"] = inq_lines["ردیف"].map(_normalize_row_key)
    row_keys = [rk for rk in inq_lines["_row_key"].unique().tolist() if rk]
    total_rows = len(row_keys)
    decided_rows = 0
    ordered_rows = 0
    pending_ids: List[str] = []

    for rk in row_keys:
        row_lines = inq_lines[inq_lines["_row_key"] == rk]
        selected_in_row = [
            line for _, line in row_lines.iterrows() if _truthy(line.get("منتخب مدیر"))
        ]
        if selected_in_row:
            decided_rows += 1
        ordered_in_row = [
            line for line in selected_in_row
            if line_has_order(str(line.get("id", "")))
        ]
        if ordered_in_row:
            ordered_rows += 1
        elif selected_in_row:
            for line in selected_in_row:
                pending_ids.append(str(line.get("id", "")))

    return {
        "total_lines": len(inq_lines),
        "total_rows": total_rows,
        "selected_lines": decided_rows,
        "lines_with_orders": ordered_rows,
        "pending_order_lines": len(pending_ids),
        "pending_row_decisions": max(0, total_rows - decided_rows),
        "pending_line_ids": pending_ids,
    }


def update_line_manager_decision(
    line_id: str,
    expert: str,
    order_number: str,
    selected: bool = True,
) -> Optional[Dict]:
    df = get_pre_invoice_lines()
    if df.empty:
        return None
    for col in LINE_HEADERS:
        if col not in df.columns:
            df[col] = None
    mask = df["id"].astype(str) == str(line_id)
    if not mask.any():
        return None
    for col in ("منتخب مدیر", "کارشناس ارجاع", "شماره دستور"):
        if col in df.columns:
            df[col] = df[col].astype(object)
    df.loc[mask, "منتخب مدیر"] = bool(selected)
    df.loc[mask, "کارشناس ارجاع"] = expert or None
    df.loc[mask, "شماره دستور"] = _normalize_order_number(order_number) or None
    _write_sheet("pre_invoice_lines", df)
    return df[mask].iloc[0].to_dict()


def inquiry_has_orders(inquiry_number: str) -> bool:
    df = get_orders()
    if df.empty:
        return False
    return str(inquiry_number) in df["شماره استعلام"].astype(str).values


def inquiry_order_count(inquiry_number: str) -> int:
    df = get_orders()
    if df.empty:
        return 0
    return int((df["شماره استعلام"].astype(str) == str(inquiry_number)).sum())


def update_inquiry_status(inquiry_number: str, status: str) -> None:
    df = get_issued_inquiries()
    if df.empty:
        return
    mask = df["شماره استعلام"].astype(str) == str(inquiry_number)
    if not mask.any():
        return
    df.loc[mask, "وضعیت"] = status
    _write_sheet("issued_inquiries", df)


def update_pre_invoice_status(
    preinvoice_id: str, status: str, reviewer: str, comment: Optional[str] = None
) -> Optional[Dict]:
    df = get_pre_invoices()
    if df.empty or "id" not in df.columns:
        return None
    df = _ensure_preinvoice_columns(df)
    mask = df["id"].astype(str) == str(preinvoice_id)
    if not mask.any():
        return None
    df["وضعیت مدیر"] = df["وضعیت مدیر"].astype(str)
    df["کامنت مدیر"] = df["کامنت مدیر"].astype(object)
    df["تاریخ بررسی"] = df["تاریخ بررسی"].astype(object)
    df["بررسی کننده"] = df["بررسی کننده"].astype(object)
    df.loc[mask, "وضعیت مدیر"] = status
    df.loc[mask, "کامنت مدیر"] = (comment or "").strip() or None
    df.loc[mask, "تاریخ بررسی"] = datetime.utcnow().isoformat()
    df.loc[mask, "بررسی کننده"] = reviewer
    _write_sheet("pre_invoices", df)
    return df[mask].iloc[0].to_dict()


def inquiry_exists(inquiry_number: str) -> bool:
    df = get_issued_inquiries()
    if df.empty:
        return False
    return inquiry_number in df["شماره استعلام"].astype(str).values


def preinvoice_duplicate(inquiry_number: str, contractor: str, preinvoice_number: str) -> bool:
    df = get_pre_invoices()
    if df.empty:
        return False
    mask = (
        (df["شماره استعلام"].astype(str) == str(inquiry_number))
        & (df["نام پیمانکار"].astype(str).str.strip() == str(contractor).strip())
        & (df["شماره پیش فاکتور"].astype(str).str.strip() == str(preinvoice_number).strip())
    )
    return bool(mask.any())


def get_product_history() -> pd.DataFrame:
    return _read_local_sheet("product_history")


def append_product_history_records(records: List[Dict[str, Any]], username: str) -> List[Dict]:
    if not records:
        return []
    now = datetime.utcnow().isoformat()
    df = get_product_history()
    rows = []
    for rec in records:
        price = float(rec.get("فی") or 0)
        if price <= 0:
            continue
        rows.append({
            "id": rec.get("id") or _new_id(),
            "کد قلم خریدنی": rec.get("کد قلم خریدنی"),
            "عنوان کالا": rec.get("عنوان کالا"),
            "فی": price,
            "تعداد": rec.get("تعداد"),
            "واحد": rec.get("واحد"),
            "نام پیمانکار": rec.get("نام پیمانکار"),
            "شهر پیمانکار": rec.get("شهر پیمانکار"),
            "شماره استعلام": rec.get("شماره استعلام"),
            "شماره خرید": rec.get("شماره خرید"),
            "تاریخ خرید": rec.get("تاریخ خرید") or now[:10],
            "created_at": now,
            "created_by": username,
        })
    if rows:
        df = pd.concat([df, pd.DataFrame(rows)], ignore_index=True)
        _write_sheet("product_history", df)
    return rows


DEFAULT_CITIES = [
    "تهران", "مشهد", "اصفهان", "شیراز", "تبریز", "کرج", "اهواز", "قم",
    "کرمانشاه", "رشت", "یزد", "اراک", "زاهدان", "همدان", "کرمان", "قزوین",
    "ساری", "بندرعباس", "خرم‌آباد", "سنندج", "گرگان", "کاشان", "بوشهر",
]


def search_cities(query: str, limit: int = 30) -> List[str]:
    names = set(DEFAULT_CITIES)
    df = get_pre_invoices()
    if not df.empty and "شهر پیمانکار" in df.columns:
        names.update(df["شهر پیمانکار"].dropna().astype(str).str.strip().tolist())
    q = query.strip().lower()
    if q:
        names = {n for n in names if n and q in n.lower()}
    return sorted(n for n in names if n)[:limit]


def search_contractors(query: str, limit: int = 20) -> List[str]:
    df = get_pre_invoices()
    if df.empty or "نام پیمانکار" not in df.columns:
        return []
    names = df["نام پیمانکار"].dropna().astype(str).unique().tolist()
    q = query.strip().lower()
    if q:
        names = [n for n in names if q in n.lower()]
    return sorted(set(names))[:limit]


def paginate_df(
    df: pd.DataFrame,
    page: int,
    page_size: int,
    search: str = "",
    search_cols: Optional[List[str]] = None,
    *,
    page_cap: int = 200,
) -> Dict:
    import math

    if search and search_cols:
        q = search.strip().lower()
        if q:
            mask = pd.Series(False, index=df.index)
            for col in search_cols:
                if col in df.columns:
                    mask |= df[col].astype(str).str.lower().str.contains(q, na=False)
            df = df[mask]
    total = len(df)
    page = max(1, page)
    page_size = max(10, min(page_size, page_cap))
    start = (page - 1) * page_size
    chunk = df.iloc[start : start + page_size]
    items = chunk.to_dict(orient="records") if not chunk.empty else []
    try:
        from services.json_util import json_safe
        items = [json_safe(item) for item in items]
    except Exception:
        pass
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": math.ceil(total / page_size) if total else 0,
    }


def get_orders() -> pd.DataFrame:
    return _read_local_sheet("orders")


def get_deliveries() -> pd.DataFrame:
    df = _read_local_sheet("deliveries")
    if df.empty:
        return df
    for col in ("انبار", "واحد", "تحویل گیرنده", "توضیحات", "عنوان کالا", "شماره تحویل", "شماره دستور", "شماره خرید"):
        if col in df.columns:
            df[col] = df[col].astype(object)
            df[col] = df[col].where(df[col].notna(), "")
    return df


def order_number_exists(order_number: str) -> bool:
    df = get_orders()
    if df.empty:
        return False
    key = _normalize_order_number(order_number)
    if not key:
        return False
    return key in df["شماره دستور"].astype(str).map(_normalize_order_number).values


def delivery_number_exists(delivery_number: str) -> bool:
    df = get_deliveries()
    if df.empty:
        return False
    return str(delivery_number) in df["شماره تحویل"].astype(str).values


def append_order(record: Dict[str, Any], username: str) -> Dict:
    now = datetime.utcnow().isoformat()
    row = {h: None for h in ORDER_HEADERS}
    row.update({
        "id": record.get("id") or _new_id(),
        "created_at": now,
        "created_by": username,
        "updated_at": now,
    })
    for k, v in record.items():
        if k in ORDER_HEADERS:
            row[k] = v
    df = get_orders()
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    _write_sheet("orders", df)
    return row


def update_order(order_id: str, updates: Dict[str, Any], username: str) -> Optional[Dict]:
    df = get_orders()
    if df.empty:
        return None
    mask = df["id"].astype(str) == str(order_id)
    if not mask.any():
        return None
    for k, v in updates.items():
        if k in df.columns:
            if df[k].dtype != object:
                df[k] = df[k].astype(object)
            df.loc[mask, k] = v
    df.loc[mask, "updated_at"] = datetime.utcnow().isoformat()
    _write_sheet("orders", df)
    return df[mask].iloc[0].to_dict()


def append_delivery(record: Dict[str, Any], username: str) -> Dict:
    now = datetime.utcnow().isoformat()
    row = {h: None for h in DELIVERY_HEADERS}
    row.update({
        "id": record.get("id") or _new_id(),
        "created_at": now,
        "created_by": username,
        "updated_at": now,
    })
    for k, v in record.items():
        if k in DELIVERY_HEADERS:
            row[k] = v
    df = get_deliveries()
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    _write_sheet("deliveries", df)
    return row


def find_delivery_by_id(delivery_id: str) -> Optional[Dict]:
    df = get_deliveries()
    if df.empty or not delivery_id:
        return None
    match = df[df["id"].astype(str) == str(delivery_id)]
    if match.empty:
        return None
    return match.iloc[0].to_dict()


def find_delivery_by_order(order_number: str) -> Optional[Dict]:
    df = get_deliveries()
    if df.empty or not order_number:
        return None
    key = _normalize_order_number(order_number)
    match = df[df["شماره دستور"].astype(str).map(_normalize_order_number) == key]
    if match.empty:
        return None
    return match.iloc[0].to_dict()


def find_order_by_number(order_number: str) -> Optional[Dict]:
    df = get_orders()
    if df.empty or not order_number:
        return None
    key = _normalize_order_number(order_number)
    match = df[df["شماره دستور"].astype(str).map(_normalize_order_number) == key]
    if match.empty:
        return None
    return match.iloc[0].to_dict()


def update_delivery(delivery_id: str, updates: Dict[str, Any], username: str) -> Optional[Dict]:
    df = get_deliveries()
    if df.empty:
        return None
    mask = df["id"].astype(str) == str(delivery_id)
    if not mask.any():
        return None
    for k, v in updates.items():
        if k in df.columns:
            if df[k].dtype != object:
                df[k] = df[k].astype(object)
            df.loc[mask, k] = v
    df.loc[mask, "updated_at"] = datetime.utcnow().isoformat()
    _write_sheet("deliveries", df)
    return df[mask].iloc[0].to_dict()


def append_notification(record: Dict[str, Any]) -> Dict:
    now = datetime.utcnow().isoformat()
    row = {
        "id": _new_id(),
        "username": record.get("username"),
        "warehouse": record.get("warehouse"),
        "عنوان": record.get("عنوان"),
        "پیام": record.get("پیام"),
        "نوع": record.get("نوع"),
        "مرجع": record.get("مرجع"),
        "خوانده شده": False,
        "created_at": now,
    }
    df = _read_local_sheet("notifications")
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    _write_sheet("notifications", df)
    return row


def get_notifications(username: str = "", unread_only: bool = False) -> List[Dict]:
    df = _read_local_sheet("notifications")
    if df.empty:
        return []
    if username:
        df = df[df["username"].astype(str) == str(username)]
    if unread_only and "خوانده شده" in df.columns:
        df = df[df["خوانده شده"].astype(str).isin(["False", "false", "0", ""]) | df["خوانده شده"].isna()]
    return df.sort_values("created_at", ascending=False).to_dict(orient="records")


def append_edit_history(record: Dict[str, Any]) -> Dict:
    now = datetime.utcnow().isoformat()
    row = {h: None for h in EDIT_HISTORY_HEADERS}
    row.update({
        "id": _new_id(),
        "created_at": now,
        "نوع موجودیت": record.get("نوع موجودیت"),
        "شناسه": str(record.get("شناسه") or ""),
        "عملیات": record.get("عملیات", "ویرایش"),
        "فیلد": record.get("فیلد"),
        "مقدار قبلی": record.get("مقدار قبلی"),
        "مقدار جدید": record.get("مقدار جدید"),
        "کاربر": record.get("کاربر"),
    })
    df = _read_local_sheet("edit_history")
    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    _write_sheet("edit_history", df)
    return row


def paginate_edit_history(
    page: int = 1,
    page_size: int = 50,
    search: str = "",
    entity_type: str = "",
    entity_id: str = "",
    *,
    page_cap: int = 200,
) -> Dict:
    df = _read_local_sheet("edit_history")
    if entity_type and "نوع موجودیت" in df.columns:
        df = df[df["نوع موجودیت"].astype(str) == str(entity_type)]
    if entity_id and "شناسه" in df.columns:
        df = df[df["شناسه"].astype(str) == str(entity_id)]
    if not df.empty:
        df = df.sort_values("created_at", ascending=False)
    search_cols = ["نوع موجودیت", "شناسه", "عملیات", "فیلد", "مقدار قبلی", "مقدار جدید", "کاربر"]
    if search.strip():
        q = search.strip().lower()
        mask = pd.Series(False, index=df.index)
        for col in search_cols:
            if col in df.columns:
                mask |= df[col].astype(str).str.lower().str.contains(q, na=False)
        df = df[mask]
    return paginate_df(df, page, page_size, page_cap=page_cap)


def mark_notification_read(notification_id: str, username: str) -> Optional[Dict]:
    df = _read_local_sheet("notifications")
    if df.empty:
        return None
    mask = (df["id"].astype(str) == str(notification_id)) & (df["username"].astype(str) == str(username))
    if not mask.any():
        return None
    df.loc[mask, "خوانده شده"] = True
    _write_sheet("notifications", df)
    return df[mask].iloc[0].to_dict()


def local_info() -> dict:
    if _use_sqlite():
        from services.db_storage import db_info
        info = db_info()
        info["inquiries"] = len(get_issued_inquiries())
        info["pre_invoices"] = len(get_pre_invoices())
        return info
    path = Path(LOCAL_EXCEL_PATH)
    return {
        "backend": "excel",
        "path": str(path),
        "exists": path.exists(),
        "size_kb": round(path.stat().st_size / 1024, 1) if path.exists() else 0,
        "last_modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat() if path.exists() else None,
        "inquiries": len(get_issued_inquiries()),
        "pre_invoices": len(get_pre_invoices()),
    }